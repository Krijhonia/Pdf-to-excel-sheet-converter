import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import pandas as pd
import tabula
import PyPDF2
import os
from pathlib import Path
import threading
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import camelot
import pdfplumber
import json

class PDFToExcelConverter:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("PDF to Excel Converter AI")
        self.root.geometry("900x700")
        self.root.configure(bg='#f0f0f0')
        
        # Variables
        self.selected_files = []
        self.output_directory = tk.StringVar()
        self.conversion_method = tk.StringVar(value="tabula")
        self.password = tk.StringVar()
        self.progress_var = tk.DoubleVar()
        self.status_var = tk.StringVar(value="Ready")
        
        # Conversion settings
        self.settings = {
            'extract_all_pages': tk.BooleanVar(value=True),
            'page_range': tk.StringVar(value="1-"),
            'multiple_tables': tk.BooleanVar(value=True),
            'format_output': tk.BooleanVar(value=True),
            'include_metadata': tk.BooleanVar(value=False)
        }
        
        self.setup_ui()
        self.load_settings()
        
    def setup_ui(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="PDF to Excel Converter AI", 
                               font=('Arial', 16, 'bold'))
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 20))
        
        # File selection section
        file_frame = ttk.LabelFrame(main_frame, text="Select PDF Files", padding="10")
        file_frame.grid(row=1, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        file_frame.columnconfigure(0, weight=1)
        
        ttk.Button(file_frame, text="Browse Files", 
                  command=self.browse_files).grid(row=0, column=0, sticky=tk.W)
        ttk.Button(file_frame, text="Clear All", 
                  command=self.clear_files).grid(row=0, column=1, sticky=tk.W, padx=(10, 0))
        ttk.Button(file_frame, text="Remove Selected", 
                  command=self.remove_selected_file).grid(row=0, column=2, sticky=tk.W, padx=(10, 0))
        
        # File list
        self.file_listbox = tk.Listbox(file_frame, height=6)
        self.file_listbox.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        
        # Scrollbar for file list
        scrollbar = ttk.Scrollbar(file_frame, orient=tk.VERTICAL, command=self.file_listbox.yview)
        scrollbar.grid(row=1, column=2, sticky=(tk.N, tk.S))
        self.file_listbox.configure(yscrollcommand=scrollbar.set)
        
        # Output directory section
        output_frame = ttk.LabelFrame(main_frame, text="Output Directory", padding="10")
        output_frame.grid(row=2, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        output_frame.columnconfigure(0, weight=1)
        
        ttk.Entry(output_frame, textvariable=self.output_directory, 
                 state='readonly').grid(row=0, column=0, sticky=(tk.W, tk.E), padx=(0, 10))
        ttk.Button(output_frame, text="Browse", 
                  command=self.browse_output_dir).grid(row=0, column=1)
        
        # Settings section
        settings_frame = ttk.LabelFrame(main_frame, text="Conversion Settings", padding="10")
        settings_frame.grid(row=3, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        # Method selection
        ttk.Label(settings_frame, text="Extraction Method:").grid(row=0, column=0, sticky=tk.W)
        method_frame = ttk.Frame(settings_frame)
        method_frame.grid(row=0, column=1, sticky=tk.W, padx=(10, 0))
        
        ttk.Radiobutton(method_frame, text="Tabula (Fast)", 
                       variable=self.conversion_method, value="tabula").grid(row=0, column=0)
        ttk.Radiobutton(method_frame, text="Camelot (Accurate)", 
                       variable=self.conversion_method, value="camelot").grid(row=0, column=1, padx=(10, 0))
        ttk.Radiobutton(method_frame, text="PDFPlumber (Advanced)", 
                       variable=self.conversion_method, value="pdfplumber").grid(row=0, column=2, padx=(10, 0))
        
        # Password field
        ttk.Label(settings_frame, text="Password (if required):").grid(row=1, column=0, sticky=tk.W, pady=(10, 0))
        ttk.Entry(settings_frame, textvariable=self.password, 
                 show="*").grid(row=1, column=1, sticky=(tk.W, tk.E), padx=(10, 0), pady=(10, 0))
        
        # Additional options
        options_frame = ttk.Frame(settings_frame)
        options_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        
        ttk.Checkbutton(options_frame, text="Extract all pages", 
                       variable=self.settings['extract_all_pages']).grid(row=0, column=0, sticky=tk.W)
        ttk.Checkbutton(options_frame, text="Handle multiple tables", 
                       variable=self.settings['multiple_tables']).grid(row=0, column=1, sticky=tk.W, padx=(20, 0))
        ttk.Checkbutton(options_frame, text="Format output", 
                       variable=self.settings['format_output']).grid(row=1, column=0, sticky=tk.W)
        ttk.Checkbutton(options_frame, text="Include metadata", 
                       variable=self.settings['include_metadata']).grid(row=1, column=1, sticky=tk.W, padx=(20, 0))
        
        # Page range
        page_frame = ttk.Frame(settings_frame)
        page_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(10, 0))
        
        ttk.Label(page_frame, text="Page Range:").grid(row=0, column=0, sticky=tk.W)
        ttk.Entry(page_frame, textvariable=self.settings['page_range'], 
                 width=20).grid(row=0, column=1, sticky=tk.W, padx=(10, 0))
        ttk.Label(page_frame, text="(e.g., 1-5, 1,3,5, or 1- for all)").grid(row=0, column=2, sticky=tk.W, padx=(10, 0))
        
        # Conversion section
        convert_frame = ttk.LabelFrame(main_frame, text="Convert", padding="10")
        convert_frame.grid(row=4, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        
        self.convert_button = ttk.Button(convert_frame, text="Start Conversion", 
                                        command=self.start_conversion)
        self.convert_button.grid(row=0, column=0, sticky=tk.W)
        
        ttk.Button(convert_frame, text="Save Settings", 
                  command=self.save_settings).grid(row=0, column=1, sticky=tk.W, padx=(10, 0))
        ttk.Button(convert_frame, text="Load Settings", 
                  command=self.load_settings).grid(row=0, column=2, sticky=tk.W, padx=(10, 0))
        
        # Progress section
        progress_frame = ttk.LabelFrame(main_frame, text="Progress", padding="10")
        progress_frame.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(0, 10))
        progress_frame.columnconfigure(0, weight=1)
        
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, 
                                           maximum=100)
        self.progress_bar.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        
        ttk.Label(progress_frame, textvariable=self.status_var).grid(row=1, column=0, sticky=tk.W)
        
        # Log section
        log_frame = ttk.LabelFrame(main_frame, text="Conversion Log", padding="10")
        log_frame.grid(row=6, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        main_frame.rowconfigure(6, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=8, width=80)
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Add Help menu
        menubar = tk.Menu(self.root)
        helpmenu = tk.Menu(menubar, tearoff=0)
        helpmenu.add_command(label="About", command=self.show_about)
        helpmenu.add_command(label="Usage", command=self.show_usage)
        menubar.add_cascade(label="Help", menu=helpmenu)
        self.root.config(menu=menubar)
        
    def browse_files(self):
        files = filedialog.askopenfilenames(
            title="Select PDF Files",
            filetypes=[("PDF files", "*.pdf"), ("All files", "*.*")]
        )
        
        for file in files:
            if file not in self.selected_files:
                self.selected_files.append(file)
                self.file_listbox.insert(tk.END, os.path.basename(file))
        
        self.log(f"Selected {len(files)} PDF files")
        
    def clear_files(self):
        self.selected_files.clear()
        self.file_listbox.delete(0, tk.END)
        self.log("Cleared all selected files")
        
    def browse_output_dir(self):
        directory = filedialog.askdirectory(title="Select Output Directory")
        if directory:
            self.output_directory.set(directory)
            self.log(f"Output directory set to: {directory}")
            
    def log(self, message):
        timestamp = datetime.now().strftime("%H:%M:%S")
        log_message = f"[{timestamp}] {message}\n"
        self.log_text.insert(tk.END, log_message)
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
    def start_conversion(self):
        if not self.selected_files:
            messagebox.showerror("Error", "Please select at least one PDF file")
            return
            
        if not self.output_directory.get():
            messagebox.showerror("Error", "Please select an output directory")
            return
            
        self.convert_button.config(state='disabled')
        self.progress_var.set(0)
        self.status_var.set("Converting...")
        
        # Start conversion in a separate thread
        thread = threading.Thread(target=self.convert_files)
        thread.daemon = True
        thread.start()
        
    def convert_files(self):
        try:
            total_files = len(self.selected_files)
            successful_conversions = 0
            self.conversion_summary = {'success': [], 'fail': []}
            for i, pdf_file in enumerate(self.selected_files):
                self.log(f"Converting: {os.path.basename(pdf_file)}")
                try:
                    success = self.convert_single_file(pdf_file)
                    if success:
                        successful_conversions += 1
                        self.conversion_summary['success'].append(os.path.basename(pdf_file))
                        self.log(f"✓ Successfully converted: {os.path.basename(pdf_file)}")
                    else:
                        self.conversion_summary['fail'].append(os.path.basename(pdf_file))
                        self.log(f"✗ Failed to convert: {os.path.basename(pdf_file)}")
                except Exception as e:
                    self.conversion_summary['fail'].append(os.path.basename(pdf_file))
                    self.log(f"✗ Error converting {os.path.basename(pdf_file)}: {str(e)}")
                # Update progress
                progress = ((i + 1) / total_files) * 100
                self.progress_var.set(progress)
            self.log(f"Conversion complete! {successful_conversions}/{total_files} files converted successfully")
            self.status_var.set(f"Complete: {successful_conversions}/{total_files} files converted")
            self.show_conversion_summary()
        except Exception as e:
            self.log(f"Conversion error: {str(e)}")
            self.status_var.set("Error occurred during conversion")
        finally:
            self.convert_button.config(state='normal')

    def show_conversion_summary(self):
        summary = f"Success: {len(self.conversion_summary['success'])}\n" \
                  f"Failed: {len(self.conversion_summary['fail'])}\n\n"
        if self.conversion_summary['success']:
            summary += "Successful files:\n" + "\n".join(self.conversion_summary['success']) + "\n\n"
        if self.conversion_summary['fail']:
            summary += "Failed files:\n" + "\n".join(self.conversion_summary['fail'])
        messagebox.showinfo("Conversion Summary", summary)

    def convert_single_file(self, pdf_file):
        try:
            base_name = os.path.splitext(os.path.basename(pdf_file))[0]
            output_file = os.path.join(self.output_directory.get(), f"{base_name}.xlsx")
            
            # Check if PDF is password protected
            if self.is_password_protected(pdf_file):
                if not self.password.get():
                    self.log(f"Password required for {os.path.basename(pdf_file)}")
                    return False
                    
            # Extract tables based on selected method
            if self.conversion_method.get() == "tabula":
                tables = self.extract_with_tabula(pdf_file)
            elif self.conversion_method.get() == "camelot":
                tables = self.extract_with_camelot(pdf_file)
            else:
                tables = self.extract_with_pdfplumber(pdf_file)
                
            if not tables:
                self.log(f"No tables found in {os.path.basename(pdf_file)}")
                return False
                
            # Save to Excel
            self.save_to_excel(tables, output_file, pdf_file)
            return True
            
        except Exception as e:
            self.log(f"Error in convert_single_file: {str(e)}")
            return False
            
    def is_password_protected(self, pdf_file):
        try:
            with open(pdf_file, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                return pdf_reader.is_encrypted
        except:
            return False
            
    def extract_with_tabula(self, pdf_file):
        try:
            password = self.password.get() if self.password.get() else None
            pages = self.get_page_range()
            
            if self.settings['multiple_tables'].get():
                tables = tabula.read_pdf(pdf_file, pages=pages, multiple_tables=True, password=password)
            else:
                tables = [tabula.read_pdf(pdf_file, pages=pages, password=password)]
                
            return [table for table in tables if not table.empty]
            
        except Exception as e:
            self.log(f"Tabula extraction error: {str(e)}")
            return []
            
    def extract_with_camelot(self, pdf_file):
        try:
            pages = self.get_page_range()
            tables = camelot.read_pdf(pdf_file, pages=pages, password=self.password.get())
            return [table.df for table in tables]
            
        except Exception as e:
            self.log(f"Camelot extraction error: {str(e)}")
            return []
            
    def extract_with_pdfplumber(self, pdf_file):
        try:
            tables = []
            password = self.password.get() if self.password.get() else None
            
            with pdfplumber.open(pdf_file, password=password) as pdf:
                pages = self.get_page_numbers(len(pdf.pages))
                
                for page_num in pages:
                    page = pdf.pages[page_num - 1]
                    page_tables = page.extract_tables()
                    
                    for table in page_tables:
                        if table:
                            df = pd.DataFrame(table[1:], columns=table[0])
                            tables.append(df)
                            
            return tables
            
        except Exception as e:
            self.log(f"PDFPlumber extraction error: {str(e)}")
            return []
            
    def get_page_range(self):
        if self.settings['extract_all_pages'].get():
            return 'all'
        else:
            return self.settings['page_range'].get()
            
    def get_page_numbers(self, total_pages):
        if self.settings['extract_all_pages'].get():
            return list(range(1, total_pages + 1))
        
        page_range = self.settings['page_range'].get()
        pages = []
        
        try:
            if '-' in page_range:
                start, end = page_range.split('-')
                start = int(start) if start else 1
                end = int(end) if end else total_pages
                pages = list(range(start, min(end + 1, total_pages + 1)))
            elif ',' in page_range:
                pages = [int(p.strip()) for p in page_range.split(',')]
                pages = [p for p in pages if 1 <= p <= total_pages]
            else:
                page_num = int(page_range)
                if 1 <= page_num <= total_pages:
                    pages = [page_num]
                    
        except ValueError:
            self.log(f"Invalid page range: {page_range}")
            pages = [1]
            
        return pages
        
    def save_to_excel(self, tables, output_file, pdf_file):
        try:
            with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
                for i, table in enumerate(tables):
                    sheet_name = f"Table_{i+1}" if len(tables) > 1 else "Data"
                    table.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    if self.settings['format_output'].get():
                        self.format_excel_sheet(writer, sheet_name, table)
                        
                # Add metadata sheet if requested
                if self.settings['include_metadata'].get():
                    self.add_metadata_sheet(writer, pdf_file, len(tables))
                    
        except Exception as e:
            self.log(f"Excel save error: {str(e)}")
            raise
            
    def format_excel_sheet(self, writer, sheet_name, table):
        try:
            worksheet = writer.sheets[sheet_name]
            
            # Header formatting
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            
            for col_num, column_title in enumerate(table.columns, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")
                
            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                        
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
                
        except Exception as e:
            self.log(f"Formatting error: {str(e)}")
            
    def add_metadata_sheet(self, writer, pdf_file, table_count):
        try:
            metadata = {
                'Source PDF': [os.path.basename(pdf_file)],
                'Conversion Date': [datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                'Extraction Method': [self.conversion_method.get()],
                'Tables Found': [table_count],
                'File Size': [f"{os.path.getsize(pdf_file) / 1024:.2f} KB"]
            }
            
            metadata_df = pd.DataFrame(metadata)
            metadata_df.to_excel(writer, sheet_name="Metadata", index=False)
            
        except Exception as e:
            self.log(f"Metadata error: {str(e)}")
            
    def save_settings(self):
        try:
            settings = {
                'output_directory': self.output_directory.get(),
                'conversion_method': self.conversion_method.get(),
                'extract_all_pages': self.settings['extract_all_pages'].get(),
                'page_range': self.settings['page_range'].get(),
                'multiple_tables': self.settings['multiple_tables'].get(),
                'format_output': self.settings['format_output'].get(),
                'include_metadata': self.settings['include_metadata'].get()
            }
            
            with open('pdf_converter_settings.json', 'w') as f:
                json.dump(settings, f, indent=2)
                
            self.log("Settings saved successfully")
            
        except Exception as e:
            self.log(f"Error saving settings: {str(e)}")
            
    def load_settings(self):
        try:
            if os.path.exists('pdf_converter_settings.json'):
                with open('pdf_converter_settings.json', 'r') as f:
                    settings = json.load(f)
                    
                self.output_directory.set(settings.get('output_directory', ''))
                self.conversion_method.set(settings.get('conversion_method', 'tabula'))
                self.settings['extract_all_pages'].set(settings.get('extract_all_pages', True))
                self.settings['page_range'].set(settings.get('page_range', '1-'))
                self.settings['multiple_tables'].set(settings.get('multiple_tables', True))
                self.settings['format_output'].set(settings.get('format_output', True))
                self.settings['include_metadata'].set(settings.get('include_metadata', False))
                
                self.log("Settings loaded successfully")
                
        except Exception as e:
            self.log(f"Error loading settings: {str(e)}")
            
    def remove_selected_file(self):
        selected_indices = self.file_listbox.curselection()
        if not selected_indices:
            self.log("No file selected to remove.")
            return
        for index in reversed(selected_indices):
            removed_file = self.selected_files.pop(index)
            self.file_listbox.delete(index)
            self.log(f"Removed file: {os.path.basename(removed_file)}")
            
    def show_about(self):
        messagebox.showinfo(
            "About",
            "PDF to Excel Converter AI\n\nVersion 1.0\nDeveloped by Your Name\n\nConverts tables from PDF files to Excel using Tabula, Camelot, or PDFPlumber."
        )

    def show_usage(self):
        messagebox.showinfo(
            "Usage",
            "1. Click 'Browse Files' to select PDF files.\n"
            "2. (Optional) Remove files with 'Remove Selected'.\n"
            "3. Choose output directory.\n"
            "4. Adjust settings as needed.\n"
            "5. Click 'Start Conversion' to begin.\n"
            "6. View progress and logs below."
        )
        
    def run(self):
        self.root.mainloop()

if __name__ == "__main__":
    # Check for required packages
    required_packages = [
        'pandas', 'tabula-py', 'PyPDF2', 'openpyxl', 
        'camelot-py[cv]', 'pdfplumber'
    ]
    
    missing_packages = []
    
    try:
        import pandas
        import tabula
        import PyPDF2
        import openpyxl
        import camelot
        import pdfplumber
    except ImportError as e:
        missing_packages.append(str(e))
        
    if missing_packages:
        print("Missing required packages. Install them using:")
        print("pip install pandas tabula-py PyPDF2 openpyxl camelot-py[cv] pdfplumber")
        print("\nMissing packages:", missing_packages)
    else:
        app = PDFToExcelConverter()
        app.run()